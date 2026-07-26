#!/usr/bin/env python3
"""
Convert catalog XLSX -> catalog.json for Labib Center website.

Usage:
    python xlsx-to-json.py catalog-template.xlsx
        -> print JSON ke stdout

    python xlsx-to-json.py catalog-template.xlsx -o ../json/catalog.json
        -> tulis langsung ke assets/json/catalog.json

    python xlsx-to-json.py catalog-template.xlsx --dry-run
        -> validasi saja, tidak menulis output

Membutuhkan:
    Python 3.8+
    openpyxl  (install: pip install openpyxl)

Sheet pertama harus memiliki header (row 1) dengan kolom:
    id, title, category, format, price, original_price,
    description, keywords, thumbnail,
    image_1, image_2, image_3, image_4, image_5,
    drive_url
"""

import json
import sys
import argparse
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write(
        "ERROR: modul 'openpyxl' tidak ditemukan.\n"
        "Install dulu: pip install openpyxl\n"
    )
    sys.exit(2)

VALID_CATEGORIES = {'Brosur', 'Undangan', 'Banner', 'Kartu Nama', 'Poster', 'Stiker'}
VALID_FORMATS = {'CDR', 'PSD'}
IMAGE_COLS = ['image_1', 'image_2', 'image_3', 'image_4', 'image_5']
REQUIRED_COLS = [
    'id', 'title', 'category', 'format', 'price', 'description',
    'keywords', 'thumbnail', 'image_1', 'drive_url'
]
ALL_COLS = [
    'id', 'title', 'category', 'format', 'price', 'original_price',
    'description', 'keywords', 'thumbnail',
    'image_1', 'image_2', 'image_3', 'image_4', 'image_5',
    'drive_url'
]
ID_PATTERN = re.compile(r'^[a-z0-9]+(-[a-z0-9]+)*$')


def stringify(v):
    """Convert cell value to trimmed string ('' if None)."""
    if v is None:
        return ''
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def parse_int(value, field, row_num, errors):
    v = stringify(value).replace('.', '').replace(',', '')
    if not v:
        return None
    try:
        n = int(float(v))
        if n < 0:
            errors.append(f"Row {row_num}: {field} negatif ({n}).")
            return None
        return n
    except ValueError:
        errors.append(f"Row {row_num}: {field} bukan angka valid: {value!r}")
        return None


def read_rows(xlsx_path):
    """Read the first sheet as list of dict rows (using header row 1)."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    header = [stringify(c.value) for c in ws[1]]
    if not header or all(not h for h in header):
        raise SystemExit(f"ERROR: sheet '{ws.title}' tidak memiliki baris header di row 1.")

    missing = [c for c in REQUIRED_COLS if c not in header]
    if missing:
        raise SystemExit(
            f"ERROR: kolom wajib tidak ditemukan di sheet '{ws.title}': {missing}\n"
            f"Kolom terdeteksi: {[h for h in header if h]}"
        )

    rows = []
    for row_idx, row_cells in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = {}
        for col_name, cell_value in zip(header, row_cells):
            if col_name:
                row_dict[col_name] = cell_value
        # Skip fully empty rows
        if not any(stringify(row_dict.get(c)) for c in header if c):
            continue
        row_dict['_row'] = row_idx
        rows.append(row_dict)

    return rows


def convert(xlsx_path, verbose=True):
    rows = read_rows(xlsx_path)
    items = []
    errors = []
    seen_ids = set()

    for row in rows:
        i = row['_row']
        id_ = stringify(row.get('id'))
        if not id_:
            continue

        if not ID_PATTERN.match(id_):
            errors.append(f"Row {i}: id {id_!r} harus kebab-case (a-z, 0-9, tanda '-')")
        if id_ in seen_ids:
            errors.append(f"Row {i}: id {id_!r} DUPLIKAT dengan baris sebelumnya")
        seen_ids.add(id_)

        title = stringify(row.get('title'))
        category = stringify(row.get('category'))
        format_ = stringify(row.get('format')).upper()
        description = stringify(row.get('description'))
        keywords = stringify(row.get('keywords'))
        thumbnail = stringify(row.get('thumbnail'))
        drive_url = stringify(row.get('drive_url'))

        price = parse_int(row.get('price'), 'price', i, errors)
        original_price = parse_int(row.get('original_price'), 'original_price', i, errors)

        images = []
        for k in IMAGE_COLS:
            v = stringify(row.get(k))
            if v:
                images.append(v)

        if not title:
            errors.append(f"Row {i} ({id_}): title kosong")
        if category not in VALID_CATEGORIES:
            errors.append(
                f"Row {i} ({id_}): category {category!r} tidak valid. "
                f"Harus salah satu: {sorted(VALID_CATEGORIES)}"
            )
        if format_ not in VALID_FORMATS:
            errors.append(f"Row {i} ({id_}): format {format_!r} tidak valid. Harus CDR atau PSD.")
        if not description:
            errors.append(f"Row {i} ({id_}): description kosong")
        if not thumbnail:
            errors.append(f"Row {i} ({id_}): thumbnail kosong")
        if not images:
            errors.append(f"Row {i} ({id_}): minimal harus ada 1 image (image_1)")
        if not drive_url:
            errors.append(f"Row {i} ({id_}): drive_url kosong")
        if price is None:
            errors.append(f"Row {i} ({id_}): price wajib diisi (angka rupiah)")
        if original_price is not None and price is not None and original_price <= price:
            errors.append(
                f"Row {i} ({id_}): original_price ({original_price}) harus > price ({price}). "
                f"Bila tidak promo, kosongkan saja."
            )

        item = {
            'id': id_,
            'title': title,
            'category': category,
            'description': description,
            'format': format_,
            'price': price if price is not None else 0,
        }
        if original_price is not None and price is not None and original_price > price:
            item['original_price'] = original_price
        item.update({
            'thumbnail': thumbnail,
            'images': images,
            'drive_url': drive_url,
            'keywords': keywords,
        })
        items.append(item)

    if errors:
        sys.stderr.write("\n=== VALIDATION ERRORS ===\n")
        for e in errors:
            sys.stderr.write(f"  - {e}\n")
        sys.stderr.write(f"\nTotal error: {len(errors)}\n\n")
        raise SystemExit(1)

    if verbose:
        sys.stderr.write(f"OK: {len(items)} item valid.\n")
    return items


def main():
    ap = argparse.ArgumentParser(
        description="Convert catalog XLSX -> catalog.json (Labib Center).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument('input', help='Path ke file XLSX input')
    ap.add_argument('-o', '--output', help='Path output JSON (default: stdout)')
    ap.add_argument('--dry-run', action='store_true',
                    help='Validasi saja, tidak menulis output')
    args = ap.parse_args()

    items = convert(args.input)

    if args.dry_run:
        sys.stderr.write("Dry-run: tidak ada file yang ditulis.\n")
        return

    out = json.dumps(items, indent=2, ensure_ascii=False) + '\n'
    if args.output:
        Path(args.output).write_text(out, encoding='utf-8')
        sys.stderr.write(f"Wrote {len(items)} item -> {args.output}\n")
    else:
        sys.stdout.write(out)


if __name__ == '__main__':
    main()
